from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

filepath = 'Portfolio Report - Traditional IRA Enhanced.xlsx'
wb = load_workbook(filepath)

header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
highlight_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

header_font = Font(bold=True, size=12, color="FFFFFF")
title_font = Font(bold=True, size=16, color="FFFFFF")
subheader_font = Font(bold=True, size=11, color="FFFFFF")
bold_font = Font(bold=True, size=10)
regular_font = Font(size=10)

thin_border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

thick_border = Border(
    left=Side(style='medium', color="000000"),
    right=Side(style='medium', color="000000"),
    top=Side(style='medium', color="000000"),
    bottom=Side(style='medium', color="000000")
)

ws_exec = wb['Executive Summary']

ws_exec['A1'].font = title_font
ws_exec['A1'].fill = header_fill
ws_exec['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_exec.merge_cells('A1:E1')
ws_exec.row_dimensions[1].height = 30

ws_exec['A2'].font = Font(italic=True, size=10)
ws_exec['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws_exec.row_dimensions[2].height = 18

ws_exec.row_dimensions[3].height = 8

ws_exec['A4'].font = subheader_font
ws_exec['A4'].fill = subheader_fill
ws_exec['A4'].alignment = Alignment(horizontal='left', vertical='center')
ws_exec.merge_cells('A4:E4')
ws_exec.row_dimensions[4].height = 22

for col_letter in ['A', 'B', 'C']:
    cell = ws_exec[col_letter + '5']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws_exec.row_dimensions[5].height = 20

for row in range(6, 15):
    ws_exec['A' + str(row)].font = bold_font
    ws_exec['A' + str(row)].fill = metric_fill
    ws_exec['A' + str(row)].border = thin_border
    ws_exec['A' + str(row)].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    ws_exec['B' + str(row)].font = Font(bold=True, size=11)
    ws_exec['B' + str(row)].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ws_exec['B' + str(row)].border = thin_border
    ws_exec['B' + str(row)].alignment = Alignment(horizontal='right', vertical='center')
    
    ws_exec['C' + str(row)].font = regular_font
    ws_exec['C' + str(row)].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    ws_exec['C' + str(row)].border = thin_border
    ws_exec['C' + str(row)].alignment = Alignment(horizontal='left', vertical='center')
    
    ws_exec.row_dimensions[row].height = 18

ws_exec.row_dimensions[15].height = 8

ws_exec['A16'].font = subheader_font
ws_exec['A16'].fill = subheader_fill
ws_exec['A16'].alignment = Alignment(horizontal='left', vertical='center')
ws_exec.merge_cells('A16:E16')
ws_exec.row_dimensions[16].height = 22

for col_letter in ['A', 'B', 'C', 'D', 'E']:
    cell = ws_exec[col_letter + '17']
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws_exec.row_dimensions[17].height = 20

for row in range(18, 22):
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        cell = ws_exec[col_letter + str(row)]
        if row == 21:
            cell.font = bold_font
            cell.fill = highlight_fill
            cell.border = thick_border
        else:
            if col_letter == 'A':
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws_exec.row_dimensions[row].height = 18

ws_exec.column_dimensions['A'].width = 28
ws_exec.column_dimensions['B'].width = 18
ws_exec.column_dimensions['C'].width = 28
ws_exec.column_dimensions['D'].width = 15
ws_exec.column_dimensions['E'].width = 15

ws_monthly = wb['Monthly Performance']

ws_monthly['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_monthly['A1'].fill = header_fill
ws_monthly['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
ws_monthly.merge_cells('A1:M1')
ws_monthly.row_dimensions[1].height = 25

ws_monthly.row_dimensions[2].height = 8

for col in range(1, 14):
    cell = ws_monthly.cell(row=3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws_monthly.row_dimensions[3].height = 22

for row in [4, 5]:
    for col in range(1, 14):
        cell = ws_monthly.cell(row=row, column=col)
        if col == 1:
            cell.font = bold_font
            cell.fill = metric_fill
        else:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    ws_monthly.row_dimensions[row].height = 18

ws_monthly.row_dimensions[6].height = 8

ws_monthly['A7'].font = subheader_font
ws_monthly['A7'].fill = subheader_fill
ws_monthly.merge_cells('A7:M7')
ws_monthly['A7'].alignment = Alignment(horizontal='left', vertical='center')
ws_monthly.row_dimensions[7].height = 20

for row in range(8, 13):
    for col in range(1, 14):
        cell = ws_monthly.cell(row=row, column=col)
        if col == 1:
            cell.font = bold_font
            cell.fill = metric_fill
        else:
            cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    ws_monthly.row_dimensions[row].height = 18

ws_monthly.row_dimensions[13].height = 8

ws_monthly['A14'].font = subheader_font
ws_monthly['A14'].fill = subheader_fill
ws_monthly.merge_cells('A14:M14')
ws_monthly['A14'].alignment = Alignment(horizontal='left', vertical='center')
ws_monthly.row_dimensions[14].height = 20

for row in range(15, 21):
    for col in range(1, 14):
        cell = ws_monthly.cell(row=row, column=col)
        if col == 1:
            cell.font = bold_font
            cell.fill = metric_fill
        else:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    ws_monthly.row_dimensions[row].height = 18

ws_monthly.row_dimensions[21].height = 8

ws_monthly['A22'].font = subheader_font
ws_monthly['A22'].fill = subheader_fill
ws_monthly.merge_cells('A22:M22')
ws_monthly['A22'].alignment = Alignment(horizontal='left', vertical='center')
ws_monthly.row_dimensions[22].height = 20

for row in range(23, 26):
    for col in range(1, 14):
        cell = ws_monthly.cell(row=row, column=col)
        if col == 1:
            cell.font = bold_font
            cell.fill = metric_fill
        else:
            cell.fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    ws_monthly.row_dimensions[row].height = 18

ws_monthly.row_dimensions[26].height = 8

ws_monthly['A27'].font = subheader_font
ws_monthly['A27'].fill = subheader_fill
ws_monthly.merge_cells('A27:M27')
ws_monthly['A27'].alignment = Alignment(horizontal='left', vertical='center')
ws_monthly.row_dimensions[27].height = 20

for row in range(28, 31):
    for col in range(1, 14):
        cell = ws_monthly.cell(row=row, column=col)
        if col == 1:
            cell.font = bold_font
            cell.fill = metric_fill
        else:
            cell.fill = PatternFill(start_color="F1DCDB", end_color="F1DCDB", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='right', vertical='center')
    ws_monthly.row_dimensions[row].height = 18

ws_monthly.column_dimensions['A'].width = 25
for col in range(2, 14):
    ws_monthly.column_dimensions[get_column_letter(col)].width = 14

wb.save(filepath)
print("Professional formatting applied successfully!")
print("- Executive Summary: Headers, metrics, styling")
print("- Monthly Performance: Section headers with color coding")
print("- Color scheme: Professional blues and grays")
