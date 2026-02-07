from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

filepath = 'Portfolio Report - Traditional IRA Enhanced.xlsx'
wb = load_workbook(filepath)
ws = wb['Executive Summary']

# Find current last row
current_last = ws.max_row

# Define colors for ACTION ITEMS section
action_color = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
action_header_font = Font(bold=True, size=11, color="FFFFFF")
action_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

# Add spacing
spacing_row = current_last + 1
ws.row_dimensions[spacing_row].height = 8

# Section header
action_row = spacing_row + 1
ws['A' + str(action_row)].value = 'ACTION ITEMS & STRATEGY'
ws['A' + str(action_row)].font = action_header_font
ws['A' + str(action_row)].fill = action_color
ws['A' + str(action_row)].alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('A' + str(action_row) + ':E' + str(action_row))
ws.row_dimensions[action_row].height = 20

actions = [
    ('Address Cash Position', 'Negative cash balance (-$19,628) indicates active margin usage. Evaluate your leverage policy: is this risk level acceptable for your investment goals? Consider setting cash buffer thresholds.'),
    ('Maintain Diversification', 'Your balanced profit sources (dividends $8.9k + price gains $18.7k) demonstrate strong diversification. Continue monitoring the 60/40 split between passive income and capital appreciation.'),
    ('Document Trading Rules', 'Your evolution from dividend collection to swing trading shows strategy adaptation. Document entry/exit criteria, position sizing rules, and stop-loss levels for repeatable execution.'),
    ('Enhance Risk Management', 'Despite excellent returns, implement position limits and rebalancing schedules. Consider: max position size, sector concentration limits, and quarterly rebalancing reviews.'),
    ('Track Benchmark Performance', 'Consistently outperforming S&P 500 validates your strategy. Maintain monthly comparative analysis to identify when market conditions favor your approach vs passive strategies.'),
]

for idx, (title, desc) in enumerate(actions, start=action_row + 1):
    ws['A' + str(idx)].value = title
    ws['A' + str(idx)].font = Font(bold=True, size=9, color="1F4788")
    ws['A' + str(idx)].fill = action_fill
    ws['A' + str(idx)].border = border
    ws['A' + str(idx)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws['B' + str(idx)].value = desc
    ws['B' + str(idx)].font = Font(size=9)
    ws['B' + str(idx)].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ws['B' + str(idx)].border = border
    ws['B' + str(idx)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws.merge_cells('B' + str(idx) + ':E' + str(idx))
    ws.row_dimensions[idx].height = 32

wb.save(filepath)
print("Action Items & Strategy section added!")
print("- 5 actionable recommendations")
print("- Risk management, diversification, and rule documentation")
print("- Benchmark tracking and performance review guidance")
