# -*- coding: utf-8 -*-
"""
MASTER PORTFOLIO FORMATTER
Automatically routes files to appropriate formatters:
- Individual portfolios (Type A/B)
- Net Worth consolidation files
"""

import sys
from pathlib import Path
from openpyxl import load_workbook

# Import formatters
from format_net_worth import format_net_worth_file


def detect_file_type(filepath):
    """
    Detect file type: 'net_worth', 'type_a', 'type_b', or 'unknown'
    """
    try:
        wb = load_workbook(filepath)
        
        # Check for Net Worth file FIRST (before checking for Executive Summary)
        if 'Data' in wb.sheetnames:
            # Check if it has consolidated structure (S&P 500 comparison)
            ws = wb['Data']
            
            for row in range(1, min(36, ws.max_row + 1)):
                cell_value = str(ws[f'A{row}'].value or '').upper()
                if 'NET WORTH' in cell_value or ('S&P' in cell_value and 'MARKET' in cell_value):
                    return 'net_worth'
        
        # Check for Type A (has Executive Summary - but not Net Worth)
        if 'Executive Summary' in wb.sheetnames:
            # Check if this Executive Summary is for Net Worth consolidation
            try:
                ws_exec = wb['Executive Summary']
                title = str(ws_exec['A1'].value or '').upper()
                if 'NET WORTH' in title:
                    return 'net_worth'
            except:
                pass
            
            return 'type_a'
        
        # Check for Type B (single Data sheet, individual portfolio)
        if 'Data' in wb.sheetnames and len(wb.sheetnames) == 1:
            return 'type_b'
        
        return 'unknown'
    
    except Exception as e:
        print(f"  [ERROR] Could not determine file type: {str(e)}")
        return 'unknown'


def format_individual_portfolio(filepath):
    """
    Format individual portfolio file (Type A or B)
    This is a placeholder for basic formatting.
    """
    print(f"Processing individual portfolio: {filepath}")
    try:
        wb = load_workbook(filepath)
        
        # Define colors
        header_fill = "1F4788"
        subheader_fill = "4472C4"
        
        # Apply basic formatting to all sheets
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_font = Font(bold=True, size=11, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format first row as header if it contains labels
            if ws['A1'].value:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(1, col)
                    if cell.value:
                        cell.font = header_font
                        cell.fill = PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
                        cell.border = thin_border
        
        wb.save(filepath)
        print(f"[OK] Individual portfolio formatted!\n")
        return True
    
    except Exception as e:
        print(f"  [ERROR] {str(e)}")
        return False


def format_portfolio_file(filepath):
    """
    Main entry point - detects file type and routes to appropriate formatter
    """
    print(f"\n{'='*60}")
    print(f"PORTFOLIO FORMATTER v3.0")
    print(f"{'='*60}\n")
    
    if not Path(filepath).exists():
        print(f"[ERROR] File not found: {filepath}")
        return False
    
    # Detect file type
    file_type = detect_file_type(filepath)
    print(f"File detected as: {file_type.upper()}")
    
    # Route to appropriate formatter
    if file_type == 'net_worth':
        return format_net_worth_file(filepath)
    
    elif file_type in ['type_a', 'type_b']:
        return format_individual_portfolio(filepath)
    
    else:
        print(f"[ERROR] Unknown file type or format not supported")
        return False


if __name__ == "__main__":
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        success = format_portfolio_file(filepath)
        sys.exit(0 if success else 1)
    else:
        print("Usage: python format_all.py <filepath>")
        sys.exit(1)
