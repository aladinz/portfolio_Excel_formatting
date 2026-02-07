from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

filepath = 'Portfolio Report - Traditional IRA Enhanced.xlsx'
wb = load_workbook(filepath)
ws = wb['Monthly Performance']

summary_header = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
summary_font = Font(bold=True, size=10, color="FFFFFF")
summary_data_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
summary_data_font = Font(bold=True, size=10)

border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

ws.row_dimensions[32].height = 8

ws['A33'].value = '12-MONTH SUMMARY'
ws['A33'].font = summary_font
ws['A33'].fill = summary_header
ws['A33'].alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('A33:M33')
ws.row_dimensions[33].height = 20

ws['A34'].value = 'Total Profit (12 Months)'
ws['A34'].font = summary_data_font
ws['A34'].fill = summary_data_fill
ws['A34'].border = border
for col in range(2, 14):
    cell = ws.cell(row=34, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B34'].value = 24833.25
ws['B34'].number_format = '$#,##0.00'

ws['A35'].value = 'Total Dividends'
ws['A35'].font = summary_data_font
ws['A35'].fill = summary_data_fill
ws['A35'].border = border
for col in range(2, 14):
    cell = ws.cell(row=35, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B35'].value = 8949.03
ws['B35'].number_format = '$#,##0.00'

ws['A36'].value = 'Total Price Gains'
ws['A36'].font = summary_data_font
ws['A36'].fill = summary_data_fill
ws['A36'].border = border
for col in range(2, 14):
    cell = ws.cell(row=36, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B36'].value = 18748.95
ws['B36'].number_format = '$#,##0.00'

ws['A37'].value = 'Total Trades'
ws['A37'].font = summary_data_font
ws['A37'].fill = summary_data_fill
ws['A37'].border = border
for col in range(2, 14):
    cell = ws.cell(row=37, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B37'].value = 43

ws['A38'].value = 'Average Monthly Return'
ws['A38'].font = summary_data_font
ws['A38'].fill = summary_data_fill
ws['A38'].border = border
for col in range(2, 14):
    cell = ws.cell(row=38, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B38'].value = 0.44
ws['B38'].number_format = '0.00%'

ws['A39'].value = 'Portfolio Value Growth'
ws['A39'].font = summary_data_font
ws['A39'].fill = summary_data_fill
ws['A39'].border = border
for col in range(2, 14):
    cell = ws.cell(row=39, column=col)
    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    cell.border = border
    cell.alignment = Alignment(horizontal='right', vertical='center')
ws['B39'].value = 'From $219,204 to $681,841'

for row in range(34, 40):
    ws.row_dimensions[row].height = 18

wb.save(filepath)
print("Monthly Performance enhancements added!")
print("- Added 12-MONTH SUMMARY section")
print("- Summary totals: Profit, Dividends, Gains, Trades")
print("- Key metrics highlighted at bottom of sheet")
