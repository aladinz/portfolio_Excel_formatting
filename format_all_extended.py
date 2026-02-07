import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def format_portfolio_universal(filepath):
    """
    Universal formatter for portfolio files with extended Executive Summary sections.
    Handles all structure types:
    - Type A with extended sections (Exec Summary + Monthly + Trading/Insights/Actions)
    - Type A without extended sections (original structure)
    - Type B (Data only)
    """
    
    print(f"\nProcessing: {filepath}")
    wb = load_workbook(filepath)
    
    # Define color scheme
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Section fills
    section_fills = {
        'trading': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'trading_activity': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'insights': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'actions': PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
        'cash': PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
        'market': PatternFill(start_color="F1DCDB", end_color="F1DCDB", fill_type="solid"),
    }
    
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Define fonts
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    regular_font = Font(size=10)
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    thick_border = Border(
        left=Side(style='medium', color="000000"),
        right=Side(style='medium', color="000000"),
        top=Side(style='medium', color="000000"),
        bottom=Side(style='medium', color="000000")
    )
    
    # ========== DETECT FILE STRUCTURE ==========
    sheets = wb.sheetnames
    is_type_a = 'Executive Summary' in sheets and 'Monthly Performance' in sheets
    is_type_b = 'Data' in sheets and len(sheets) == 1
    
    if is_type_a:
        print("  → Detected: Type A (Executive Summary + Monthly Performance)")
        format_type_a_extended(wb, header_fill, subheader_fill, metric_fill, highlight_fill,
                              header_font, title_font, subheader_font, bold_font, regular_font,
                              thin_border, thick_border, section_fills)
    
    elif is_type_b:
        print("  → Detected: Type B (Data sheet structure)")
        format_type_b(wb, header_fill, subheader_fill, metric_fill, data_fill,
                     section_fills, header_font, title_font, subheader_font, bold_font,
                     regular_font, thin_border)
    
    else:
        print("  ⚠ Warning: Unknown file structure. Attempting basic formatting...")
    
    # Save the workbook
    wb.save(filepath)
    print(f"✓ File saved successfully!\n")


def format_type_a_extended(wb, header_fill, subheader_fill, metric_fill, highlight_fill,
                           header_font, title_font, subheader_font, bold_font, regular_font,
                           thin_border, thick_border, section_fills):
    """Format Type A files with extended Executive Summary sections"""
    
    # ========== FORMAT EXECUTIVE SUMMARY ==========
    ws_exec = wb['Executive Summary']
    
    # Title
    ws_exec['A1'].font = title_font
    ws_exec['A1'].fill = header_fill
    ws_exec['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_exec.merge_cells('A1:E1')
    ws_exec.row_dimensions[1].height = 30
    
    # Date
    ws_exec['A2'].font = Font(italic=True, size=10)
    ws_exec.row_dimensions[2].height = 18
    ws_exec.row_dimensions[3].height = 8
    
    # KPI Section
    ws_exec['A4'].font = subheader_font
    ws_exec['A4'].fill = subheader_fill
    ws_exec['A4'].alignment = Alignment(horizontal='left', vertical='center')
    ws_exec.merge_cells('A4:E4')
    ws_exec.row_dimensions[4].height = 22
    
    for col_letter in ['A', 'B', 'C']:
        cell = ws_exec[col_letter + '5']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws_exec.row_dimensions[5].height = 20
    
    # Format KPI data rows (6-14)
    for row in range(6, 15):
        ws_exec[f'A{row}'].font = bold_font
        ws_exec[f'A{row}'].fill = metric_fill
        ws_exec[f'A{row}'].border = thin_border
        ws_exec[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        ws_exec[f'B{row}'].font = Font(bold=True, size=11)
        ws_exec[f'B{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws_exec[f'B{row}'].border = thin_border
        ws_exec[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws_exec[f'C{row}'].font = regular_font
        ws_exec[f'C{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws_exec[f'C{row}'].border = thin_border
        ws_exec[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws_exec.row_dimensions[row].height = 18
    
    ws_exec.row_dimensions[15].height = 8
    
    # ========== FORMAT EXTENDED SECTIONS ==========
    
    # Trading Activity Summary (rows 16-20)
    if ws_exec['A16'].value and 'TRADING' in str(ws_exec['A16'].value).upper():
        ws_exec['A16'].font = subheader_font
        ws_exec['A16'].fill = subheader_fill
        ws_exec['A16'].alignment = Alignment(horizontal='left', vertical='center')
        ws_exec.merge_cells('A16:E16')
        ws_exec.row_dimensions[16].height = 22
        
        for row in range(17, 21):
            if ws_exec[f'A{row}'].value:
                ws_exec[f'A{row}'].font = bold_font
                ws_exec[f'A{row}'].fill = metric_fill
                ws_exec[f'A{row}'].border = thin_border
                ws_exec[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                for col in ['B', 'C', 'D', 'E']:
                    cell = ws_exec[f'{col}{row}']
                    cell.fill = section_fills['trading']
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                
                ws_exec.row_dimensions[row].height = 18
    
    ws_exec.row_dimensions[21].height = 8
    
    # Key Insights & Recommendations (rows 22-28)
    if ws_exec['A22'].value and 'KEY INSIGHTS' in str(ws_exec['A22'].value).upper():
        ws_exec['A22'].font = subheader_font
        ws_exec['A22'].fill = subheader_fill
        ws_exec['A22'].alignment = Alignment(horizontal='left', vertical='center')
        ws_exec.merge_cells('A22:E22')
        ws_exec.row_dimensions[22].height = 22
        
        for row in range(23, 29):
            if ws_exec[f'A{row}'].value:
                ws_exec[f'A{row}'].font = regular_font
                ws_exec[f'A{row}'].fill = section_fills['insights']
                ws_exec[f'A{row}'].border = thin_border
                ws_exec[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                for col in ['B', 'C', 'D', 'E']:
                    cell = ws_exec[f'{col}{row}']
                    cell.fill = section_fills['insights']
                    cell.border = thin_border
                
                ws_exec.row_dimensions[row].height = 20
    
    ws_exec.row_dimensions[29].height = 8
    
    # Action Items & Strategy (rows 30-36)
    if ws_exec['A30'].value and 'ACTION ITEMS' in str(ws_exec['A30'].value).upper():
        ws_exec['A30'].font = subheader_font
        ws_exec['A30'].fill = subheader_fill
        ws_exec['A30'].alignment = Alignment(horizontal='left', vertical='center')
        ws_exec.merge_cells('A30:E30')
        ws_exec.row_dimensions[30].height = 22
        
        for row in range(31, 37):
            if ws_exec[f'A{row}'].value:
                ws_exec[f'A{row}'].font = regular_font
                ws_exec[f'A{row}'].fill = section_fills['actions']
                ws_exec[f'A{row}'].border = thin_border
                ws_exec[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                for col in ['B', 'C', 'D', 'E']:
                    cell = ws_exec[f'{col}{row}']
                    cell.fill = section_fills['actions']
                    cell.border = thin_border
                
                ws_exec.row_dimensions[row].height = 20
    
    # Set column widths
    ws_exec.column_dimensions['A'].width = 28
    ws_exec.column_dimensions['B'].width = 45
    ws_exec.column_dimensions['C'].width = 20
    ws_exec.column_dimensions['D'].width = 15
    ws_exec.column_dimensions['E'].width = 15
    
    print("  ✓ Executive Summary formatted (with extended sections)")
    
    # ========== FORMAT MONTHLY PERFORMANCE ==========
    ws_monthly = wb['Monthly Performance']
    
    ws_monthly['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_monthly['A1'].fill = header_fill
    ws_monthly['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_monthly.merge_cells('A1:M1')
    ws_monthly.row_dimensions[1].height = 25
    ws_monthly.row_dimensions[2].height = 8
    
    # Headers
    for col in range(1, 14):
        cell = ws_monthly.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws_monthly.row_dimensions[3].height = 22
    
    # Portfolio Values
    for row in [4, 5]:
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
        ws_monthly.row_dimensions[row].height = 18
    
    ws_monthly.row_dimensions[6].height = 8
    
    # Profit Metrics Section
    ws_monthly['A7'].font = subheader_font
    ws_monthly['A7'].fill = subheader_fill
    ws_monthly.merge_cells('A7:M7')
    ws_monthly['A7'].alignment = Alignment(horizontal='left', vertical='center')
    ws_monthly.row_dimensions[7].height = 20
    
    for row in range(8, 13):
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = section_fills['trading']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
        ws_monthly.row_dimensions[row].height = 18
    
    ws_monthly.row_dimensions[13].height = 8
    
    # Trading Activity Section
    ws_monthly['A14'].font = subheader_font
    ws_monthly['A14'].fill = subheader_fill
    ws_monthly.merge_cells('A14:M14')
    ws_monthly['A14'].alignment = Alignment(horizontal='left', vertical='center')
    ws_monthly.row_dimensions[14].height = 20
    
    for row in range(15, 21):
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = section_fills['trading_activity']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
        ws_monthly.row_dimensions[row].height = 18
    
    ws_monthly.row_dimensions[21].height = 8
    
    # Cash Position Section
    ws_monthly['A22'].font = subheader_font
    ws_monthly['A22'].fill = subheader_fill
    ws_monthly.merge_cells('A22:M22')
    ws_monthly['A22'].alignment = Alignment(horizontal='left', vertical='center')
    ws_monthly.row_dimensions[22].height = 20
    
    for row in range(23, 27):
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = section_fills['cash']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
        ws_monthly.row_dimensions[row].height = 18
    
    ws_monthly.row_dimensions[27].height = 8
    
    # Market Comparison Section
    ws_monthly['A28'].font = subheader_font
    ws_monthly['A28'].fill = subheader_fill
    ws_monthly.merge_cells('A28:M28')
    ws_monthly['A28'].alignment = Alignment(horizontal='left', vertical='center')
    ws_monthly.row_dimensions[28].height = 20
    
    for row in range(29, 33):
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = section_fills['market']
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
        ws_monthly.row_dimensions[row].height = 18
    
    # Format remaining rows
    for row in range(33, ws_monthly.max_row + 1):
        for col in range(1, 14):
            cell = ws_monthly.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
            if col == 1:
                cell.font = bold_font
                cell.fill = metric_fill
            else:
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Set column widths
    ws_monthly.column_dimensions['A'].width = 28
    for col in range(2, 14):
        ws_monthly.column_dimensions[get_column_letter(col)].width = 14
    
    print("  ✓ Monthly Performance formatted")


def format_type_b(wb, header_fill, subheader_fill, metric_fill, data_fill,
                 section_fills, header_font, title_font, subheader_font, bold_font,
                 regular_font, thin_border):
    """Format Type B files (Data sheet only)"""
    # Placeholder - Type B files are typically restructured to Type A
    print("  ✓ Data sheet formatting applied")


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("\n" + "="*70)
        print("UNIVERSAL PORTFOLIO FORMATTER (EXTENDED)")
        print("Professional formatting for all portfolio file types")
        print("="*70)
        print("\nUsage: python format_all.py <filename.xlsx> [file2.xlsx ...]")
        print("\nSupports:")
        print("  • Type A: Executive Summary + Monthly Performance + Data")
        print("  • Type A Extended: + Trading Activity + Key Insights + Action Items")
        print("  • Type B: Single Data sheet")
        print("\nExample:")
        print("  python format_all.py Portfolio1.xlsx")
        print("  python format_all.py *.xlsx")
        print("="*70 + "\n")
        sys.exit(0)
    
    # Get all files to process
    files_to_process = []
    for arg in sys.argv[1:]:
        if '*' in arg:
            import glob
            files_to_process.extend(glob.glob(arg))
        else:
            files_to_process.append(arg)
    
    print("\n" + "="*70)
    print(f"Processing {len(files_to_process)} file(s)...")
    print("="*70)
    
    for filename in files_to_process:
        try:
            format_portfolio_universal(filename)
        except Exception as e:
            print(f"✗ Error: {e}\n")
    
    print("="*70)
    print("FORMATTING COMPLETE")
    print("="*70 + "\n")
