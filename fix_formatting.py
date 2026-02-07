from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import glob

def format_portfolio_complete(filepath):
    """Complete professional formatting for Type A portfolios"""
    
    print(f"\nFormatting: {filepath}")
    wb = load_workbook(filepath)
    
    # Define color scheme
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Sections
    profit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    trading_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    cash_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    market_fill = PatternFill(start_color="F1DCDB", end_color="F1DCDB", fill_type="solid")
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Define fonts
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    regular_font = Font(size=10)
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    thick_border = Border(
        left=Side(style='medium', color="000000"),
        right=Side(style='medium', color="000000"),
        top=Side(style='medium', color="000000"),
        bottom=Side(style='medium', color="000000")
    )
    
    # ========== FORMAT EXECUTIVE SUMMARY ==========
    if 'Executive Summary' in wb.sheetnames:
        ws_exec = wb['Executive Summary']
        
        # Title row
        ws_exec['A1'].font = title_font
        ws_exec['A1'].fill = header_fill
        ws_exec['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_exec.merge_cells('A1:E1')
        ws_exec.row_dimensions[1].height = 30
        
        # Date row
        ws_exec.row_dimensions[2].height = 18
        ws_exec.row_dimensions[3].height = 8
        
        # KPI Section header
        ws_exec['A4'].font = subheader_font
        ws_exec['A4'].fill = subheader_fill
        ws_exec['A4'].alignment = Alignment(horizontal='left', vertical='center')
        ws_exec.merge_cells('A4:E4')
        ws_exec.row_dimensions[4].height = 22
        
        # KPI header row
        for col_letter in ['A', 'B', 'C']:
            cell = ws_exec[col_letter + '5']
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws_exec.row_dimensions[5].height = 20
        
        # Format data rows (6-14)
        for row in range(6, 15):
            ws_exec[f'A{row}'].font = bold_font
            ws_exec[f'A{row}'].fill = metric_fill
            ws_exec[f'A{row}'].border = thin_border
            ws_exec[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws_exec[f'B{row}'].font = Font(bold=True, size=11)
            ws_exec[f'B{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            ws_exec[f'B{row}'].border = thin_border
            ws_exec[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
            
            ws_exec[f'C{row}'].font = regular_font
            ws_exec[f'C{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            ws_exec[f'C{row}'].border = thin_border
            ws_exec[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws_exec.row_dimensions[row].height = 18
        
        # Set column widths
        ws_exec.column_dimensions['A'].width = 28
        ws_exec.column_dimensions['B'].width = 18
        ws_exec.column_dimensions['C'].width = 28
        
        print("  ✓ Executive Summary formatted")
    
    # ========== FORMAT MONTHLY PERFORMANCE ==========
    if 'Monthly Performance' in wb.sheetnames:
        ws_monthly = wb['Monthly Performance']
        
        # Title
        ws_monthly['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws_monthly['A1'].fill = header_fill
        ws_monthly['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_monthly.merge_cells('A1:M1')
        ws_monthly.row_dimensions[1].height = 25
        ws_monthly.row_dimensions[2].height = 8
        
        # Headers row 3
        for col in range(1, 14):
            cell = ws_monthly.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
        ws_monthly.row_dimensions[3].height = 22
        
        # Portfolio values rows 4-5
        for row in [4, 5]:
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws_monthly.row_dimensions[row].height = 18
        
        ws_monthly.row_dimensions[6].height = 8
        
        # PROFIT METRICS Section (rows 7-13)
        ws_monthly['A7'].font = subheader_font
        ws_monthly['A7'].fill = subheader_fill
        ws_monthly.merge_cells('A7:M7')
        ws_monthly['A7'].alignment = Alignment(horizontal='left', vertical='center')
        ws_monthly.row_dimensions[7].height = 20
        
        for row in range(8, 13):
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = profit_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws_monthly.row_dimensions[row].height = 18
        
        ws_monthly.row_dimensions[13].height = 8
        
        # TRADING ACTIVITY Section (rows 14-20)
        ws_monthly['A14'].font = subheader_font
        ws_monthly['A14'].fill = subheader_fill
        ws_monthly.merge_cells('A14:M14')
        ws_monthly['A14'].alignment = Alignment(horizontal='left', vertical='center')
        ws_monthly.row_dimensions[14].height = 20
        
        for row in range(15, 21):
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = trading_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws_monthly.row_dimensions[row].height = 18
        
        ws_monthly.row_dimensions[21].height = 8
        
        # CASH POSITION Section (rows 22-26)
        ws_monthly['A22'].font = subheader_font
        ws_monthly['A22'].fill = subheader_fill
        ws_monthly.merge_cells('A22:M22')
        ws_monthly['A22'].alignment = Alignment(horizontal='left', vertical='center')
        ws_monthly.row_dimensions[22].height = 20
        
        for row in range(23, 27):
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = cash_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws_monthly.row_dimensions[row].height = 18
        
        ws_monthly.row_dimensions[27].height = 8
        
        # MARKET COMPARISON Section (rows 28-32)
        ws_monthly['A28'].font = subheader_font
        ws_monthly['A28'].fill = subheader_fill
        ws_monthly.merge_cells('A28:M28')
        ws_monthly['A28'].alignment = Alignment(horizontal='left', vertical='center')
        ws_monthly.row_dimensions[28].height = 20
        
        for row in range(29, 33):
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = market_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
            ws_monthly.row_dimensions[row].height = 18
        
        # Format remaining rows
        for row in range(33, ws_monthly.max_row + 1):
            for col in range(1, 14):
                cell = ws_monthly.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if col == 1:
                    cell.font = bold_font
                    cell.fill = metric_fill
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        # Set column widths
        ws_monthly.column_dimensions['A'].width = 28
        for col in range(2, 14):
            ws_monthly.column_dimensions[get_column_letter(col)].width = 14
        
        print("  ✓ Monthly Performance formatted")
    
    # Save the workbook
    wb.save(filepath)
    print(f"  ✓ File saved successfully!\n")


# Process Investment, Rollover IRA, and Roth IRA files
files_to_format = [
    'Portfolio report_Investment_07.02.2026.xlsx',
    'Portfolio report_Rollover IRA_07.02.2026.xlsx',
    'Portfolio report_Roth IRA_07.02.2026.xlsx'
]

print('\n' + '='*70)
print('APPLYING COMPLETE FORMATTING TO RESTRUCTURED FILES')
print('='*70)

for filepath in files_to_format:
    if os.path.exists(filepath):
        try:
            format_portfolio_complete(filepath)
        except Exception as e:
            print(f"  ✗ Error: {e}\n")
    else:
        print(f"  ✗ File not found: {filepath}\n")

print('='*70)
print('FORMATTING COMPLETE - All restructured files now professionally formatted')
print('='*70 + '\n')
