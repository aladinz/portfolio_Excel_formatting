# -*- coding: utf-8 -*-
"""
NET WORTH CONSOLIDATION FORMATTER
Formats the consolidated "My Net Worth" portfolio file with:
- Professional styling and colors
- Executive Summary with consolidation metrics
- Comparison charts (Portfolio vs S&P 500)
- Performance metrics across all portfolios
- Win/Loss indicators
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from datetime import datetime

def format_net_worth_file(filepath):
    """
    Format Net Worth consolidation file with professional styling and analytics.
    """
    print(f"\nProcessing: {filepath}")
    
    try:
        wb = load_workbook(filepath)
        
        # Check if it's a Net Worth file
        if 'Data' not in wb.sheetnames:
            print("  [ERROR] Net Worth file must have 'Data' sheet")
            return False
        
        # Define color scheme
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        comparison_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        # Fonts
        header_font = Font(bold=True, size=11, color="FFFFFF")
        title_font = Font(bold=True, size=14, color="FFFFFF")
        subheader_font = Font(bold=True, size=12, color="FFFFFF")
        bold_font = Font(bold=True, size=11)
        regular_font = Font(size=10)
        
        # Borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format existing Data sheet
        try:
            format_data_sheet(wb, header_fill, subheader_fill, metric_fill, data_fill, 
                             header_font, title_font, bold_font, regular_font, thin_border)
        except Exception as e:
            print(f"  [ERROR] Data sheet formatting: {str(e)}")
        
        # Create Executive Summary sheet
        try:
            create_executive_summary(wb, header_fill, subheader_fill, metric_fill,
                                    header_font, title_font, bold_font, regular_font, thin_border)
        except Exception as e:
            print(f"  [ERROR] Executive Summary creation: {str(e)}")
        
        # Add charts to Executive Summary
        try:
            add_comparison_charts(wb)
        except Exception as e:
            print(f"  [ERROR] Chart creation: {str(e)}")
        
        # Save the workbook
        wb.save(filepath)
        print(f"[OK] Net Worth file formatted successfully!\n")
        return True
        
    except Exception as e:
        print(f"  [ERROR] {str(e)}")
        return False


def format_data_sheet(wb, header_fill, subheader_fill, metric_fill, data_fill,
                     header_font, title_font, bold_font, regular_font, thin_border):
    """Format the Data sheet with professional styling."""
    ws = wb['Data']
    
    # Format title row
    ws['A1'].font = title_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 25
    
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 8
    
    # Format month headers (row 4)
    for col in range(1, 14):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[4].height = 20
    
    # Format data rows with section headers
    section_colors = {
        'portfolio': PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
        'profit': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'turnover': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'trading': PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        'cash': PatternFill(start_color="F1DCDB", end_color="F1DCDB", fill_type="solid"),
        'benchmark': PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"),
    }
    
    current_section = None
    
    # Format all data rows (5-35)
    for row in range(5, 36):
        label_cell = ws[f'A{row}']
        label = str(label_cell.value or '').upper() if label_cell.value else ''
        
        # Determine section
        if 'PORTFOLIO VALUE' in label or 'AT THE' in label or 'CHANGE' in label:
            current_section = 'portfolio'
        elif 'PROFIT' in label or 'TAX' in label or 'COMMISSION' in label or 'DIVIDEND' in label:
            current_section = 'profit'
        elif 'TURNOVER' in label or 'PURCHASE' in label or 'SALE' in label:
            current_section = 'turnover'
        elif 'TRADE' in label or 'BUY' in label or 'SELL' in label:
            current_section = 'trading'
        elif 'CASH' in label or 'DEPOSIT' in label or 'WITHDRAW' in label:
            current_section = 'cash'
        elif 'S&P' in label or 'MARKET' in label:
            current_section = 'benchmark'
        
        # Apply formatting to row
        section_fill = section_colors.get(current_section, data_fill)
        
        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            
            if col == 1:  # Label column
                cell.font = Font(bold=True, size=11, color="000000")  # Dark/black text
                cell.fill = metric_fill
            else:
                cell.fill = section_fill
            
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Format numbers
            if cell.value and col > 1 and isinstance(cell.value, (int, float)):
                if 'PROFIT' in label and '%' in label:
                    cell.number_format = '0.0"%"'
                else:
                    cell.number_format = '#,##0.00'
        
        ws.row_dimensions[row].height = 18
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 13


def create_executive_summary(wb, header_fill, subheader_fill, metric_fill,
                           header_font, title_font, bold_font, regular_font, thin_border):
    """Create Executive Summary sheet for Net Worth consolidation."""
    
    # Enhanced color scheme
    subheader_font_local = Font(bold=True, size=12, color="FFFFFF")
    accent_color_green = "70AD47"  # Green for gains
    accent_color_red = "C5504F"    # Red for losses
    metric_value_font = Font(bold=True, size=13, color="FFFFFF")
    section_header_fill = PatternFill(start_color="2E5C8A", end_color="2E5C8A", fill_type="solid")
    metric_label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    metric_value_fill_positive = PatternFill(start_color=accent_color_green, end_color=accent_color_green, fill_type="solid")
    metric_value_fill_neutral = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Create new sheet
    ws = wb.create_sheet('Executive Summary', 0)
    
    # Title section with enhanced styling
    ws['A1'].value = "NET WORTH CONSOLIDATION - EXECUTIVE SUMMARY"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:F1')
    ws.row_dimensions[1].height = 30
    
    # Subtitle with date
    ws['A2'].value = f"Performance Report | As of {datetime.now().strftime('%B %d, %Y')}"
    ws['A2'].font = Font(italic=True, size=10, color="666666")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A2:F2')
    ws.row_dimensions[2].height = 18
    
    ws.row_dimensions[3].height = 6
    
    # Get data from Data sheet for metrics
    data_ws = wb['Data']
    
    # Extract key metrics
    metrics = extract_net_worth_metrics(data_ws)
    
    # Portfolio Performance Section
    row = 4
    ws[f'A{row}'].value = "💼 PORTFOLIO PERFORMANCE"
    ws[f'A{row}'].font = subheader_font_local
    ws[f'A{row}'].fill = section_header_fill
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Key metrics display with enhanced formatting
    perf_metrics = [
        ("Total Net Worth (Latest)", f"${metrics.get('latest_value', 0):,.2f}", metric_value_fill_neutral),
        ("Total Profit/Loss (YTD)", f"${metrics.get('total_profit', 0):,.2f}", 
         metric_value_fill_positive if metrics.get('total_profit', 0) >= 0 else PatternFill(start_color=accent_color_red, end_color=accent_color_red, fill_type="solid")),
        ("Profit %", f"{metrics.get('profit_pct', 0):.2f}%", metric_value_fill_neutral),
        ("Largest Monthly Gain", f"${metrics.get('largest_gain', 0):,.2f}", metric_value_fill_positive),
        ("Largest Monthly Loss", f"${metrics.get('largest_loss', 0):,.2f}", 
         PatternFill(start_color=accent_color_red, end_color=accent_color_red, fill_type="solid")),
    ]
    
    for metric_name, metric_value, value_fill in perf_metrics:
        # Label cell
        ws[f'A{row}'].value = metric_name
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = metric_label_fill
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Value cell with color coding
        ws[f'B{row}'].value = metric_value
        ws[f'B{row}'].font = metric_value_font
        ws[f'B{row}'].fill = value_fill
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'B{row}:F{row}')
        
        ws.row_dimensions[row].height = 24
        row += 1
    
    row += 1
    
    # S&P 500 Comparison Section
    ws[f'A{row}'].value = "📊 VS. S&P 500 BENCHMARK"
    ws[f'A{row}'].font = subheader_font_local
    ws[f'A{row}'].fill = section_header_fill
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 22
    row += 1
    
    comparison_metrics = [
        ("Your Portfolio Gain", f"${metrics.get('portfolio_gain', 0):,.2f}", metric_value_fill_positive),
        ("S&P 500 Gain (Est.)", f"${metrics.get('sp500_gain', 0):,.2f}", metric_value_fill_neutral),
        ("Outperformance", f"${metrics.get('outperformance', 0):,.2f}", 
         metric_value_fill_positive if metrics.get('outperformance', 0) >= 0 else PatternFill(start_color=accent_color_red, end_color=accent_color_red, fill_type="solid")),
        ("Your Portfolio %", f"{metrics.get('portfolio_pct', 0):.2f}%", metric_value_fill_positive),
        ("S&P 500 %", f"{metrics.get('sp500_pct', 0):.2f}%", metric_value_fill_neutral),
    ]
    
    for metric_name, metric_value, value_fill in comparison_metrics:
        # Label cell
        ws[f'A{row}'].value = metric_name
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = metric_label_fill
        ws[f'A{row}'].border = thin_border
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        # Value cell with color coding
        ws[f'B{row}'].value = metric_value
        ws[f'B{row}'].font = metric_value_font
        ws[f'B{row}'].fill = value_fill
        ws[f'B{row}'].border = thin_border
        ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws.merge_cells(f'B{row}:F{row}')
        
        ws.row_dimensions[row].height = 24
        row += 1
    
    # Column widths for enhanced layout
    ws.column_dimensions['A'].width = 28
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18


def extract_net_worth_metrics(ws):
    """Extract key metrics from Data sheet."""
    metrics = {
        'latest_value': 0,
        'total_profit': 0,
        'profit_pct': 0,
        'largest_gain': 0,
        'largest_loss': 0,
        'portfolio_gain': 0,
        'sp500_gain': 0,
        'outperformance': 0,
        'portfolio_pct': 0,
        'sp500_pct': 0,
    }
    
    try:
        # Get latest portfolio value (last month, column M, row 5)
        latest = ws['M5'].value
        if latest:
            metrics['latest_value'] = float(latest) if isinstance(latest, (int, float)) else 0
        
        # Get INITIAL portfolio value (first month, column B, row 5)
        initial = ws['B5'].value
        initial_value = float(initial) if isinstance(initial, (int, float)) else 0
        
        # Calculate Total YTD Profit as SUM of all monthly profits (row 10)
        total_profit = 0
        for col in range(2, 14):  # B through M (12 months)
            val = ws.cell(10, col).value
            if val and isinstance(val, (int, float)):
                total_profit += float(val)
        
        metrics['total_profit'] = total_profit
        
        # Calculate profit percentage
        if initial_value != 0:
            metrics['profit_pct'] = (total_profit / initial_value) * 100
        
        # Find largest gain and loss from profit row (row 10)
        profits = []
        for col in range(2, 14):  # B through M
            val = ws.cell(10, col).value
            if val and isinstance(val, (int, float)):
                profits.append(float(val))
        
        if profits:
            metrics['largest_gain'] = max([p for p in profits if p >= 0], default=0)
            metrics['largest_loss'] = min([p for p in profits if p < 0], default=0)
        
        # Calculate Portfolio Gain (difference between latest and initial values)
        metrics['portfolio_gain'] = metrics['latest_value'] - initial_value
        
        # Get S&P 500 values (row 34)
        sp500_initial = ws['B34'].value
        sp500_initial = float(sp500_initial) if isinstance(sp500_initial, (int, float)) else 0
        
        sp500_latest = ws['M34'].value
        sp500_latest_val = float(sp500_latest) if isinstance(sp500_latest, (int, float)) else 0
        
        # Calculate S&P 500 Gain
        metrics['sp500_gain'] = sp500_latest_val - sp500_initial
        
        # Calculate outperformance
        metrics['outperformance'] = metrics['portfolio_gain'] - metrics['sp500_gain']
        
        # Calculate percentages
        if initial_value != 0:
            metrics['portfolio_pct'] = (metrics['portfolio_gain'] / initial_value) * 100
        
        if sp500_initial != 0:
            metrics['sp500_pct'] = (metrics['sp500_gain'] / sp500_initial) * 100
        
    except Exception as e:
        pass
    
    return metrics


def add_comparison_charts(wb):
    """Add Net Worth vs S&P 500 comparison bar chart with professional styling."""
    try:
        if 'Executive Summary' not in wb.sheetnames:
            return
        
        ws_exec = wb['Executive Summary']
        ws_data = wb['Data']
        
        # Get months and data
        months = []
        portfolio_values = []
        sp500_values = []
        
        for col in range(2, 14):  # B through M (12 months)
            col_letter = get_column_letter(col)
            
            # Month
            month = ws_data[f'{col_letter}4'].value
            if month:
                months.append(str(month))
            
            # Portfolio value (row 5)
            pval = ws_data[f'{col_letter}5'].value
            if pval:
                try:
                    portfolio_values.append(float(pval) if isinstance(pval, (int, float)) else 0)
                except:
                    portfolio_values.append(0)
            
            # S&P 500 value (row 34)
            sp_val = ws_data[f'{col_letter}34'].value
            if sp_val:
                try:
                    sp500_values.append(float(sp_val) if isinstance(sp_val, (int, float)) else 0)
                except:
                    sp500_values.append(0)
        
        if months and portfolio_values and sp500_values:
            # Add data to reference area (hidden columns)
            ws_exec['H2'].value = "Month"
            ws_exec['I2'].value = "Your Portfolio"
            ws_exec['J2'].value = "S&P 500"
            
            for i, (month, pval, sp_val) in enumerate(zip(months, portfolio_values, sp500_values)):
                ws_exec[f'H{3+i}'].value = month
                ws_exec[f'I{3+i}'].value = pval
                ws_exec[f'J{3+i}'].value = sp_val
            
            # Create single professional bar chart
            bar_chart = BarChart()
            bar_chart.type = "col"  # Column (vertical bars)
            bar_chart.title = "Portfolio Value vs S&P 500 (Monthly Track)"
            bar_chart.y_axis.title = "Value ($)"
            bar_chart.height = 12
            bar_chart.width = 20
            bar_chart.legend.position = 'b'  # Bottom legend
            bar_chart.legend.horizontalAnchor = "center"
            
            # Add data series
            data_portfolio = Reference(ws_exec, min_col=9, min_row=2, max_row=2+len(portfolio_values))
            data_sp500 = Reference(ws_exec, min_col=10, min_row=2, max_row=2+len(sp500_values))
            categories = Reference(ws_exec, min_col=8, min_row=3, max_row=2+len(months))
            
            bar_chart.add_data(data_portfolio, titles_from_data=True)
            bar_chart.add_data(data_sp500, titles_from_data=True)
            bar_chart.set_categories(categories)
            
            # Color the series
            bar_chart.series[0].graphicalProperties.solidFill = "1F4788"  # Dark blue - Your portfolio
            bar_chart.series[1].graphicalProperties.solidFill = "70AD47"  # Green - S&P 500
            
            # Remove data labels for clean look
            from openpyxl.chart.label import DataLabelList
            bar_chart.dataLabels = DataLabelList()
            bar_chart.dataLabels.showVal = False
            
            ws_exec.add_chart(bar_chart, "A18")
            
    except Exception as e:
        pass


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        format_net_worth_file(filepath)
    else:
        print("Usage: python format_net_worth.py <filepath>")
