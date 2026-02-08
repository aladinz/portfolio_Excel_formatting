# -*- coding: utf-8 -*-
import sys
import os
import io

# Fix Unicode output for Windows console
if sys.platform == 'win32':
    # Reconfigure stdout to use UTF-8 encoding
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def restructure_type_b_to_type_a(filepath):
    """
    Convert Type B (single Data sheet) to Type A (Executive Summary + Monthly Performance)
    Automatically extracts data, calculates KPIs, creates structured sheets, applies formatting
    """
    
    print(f"\nRestructuring: {filepath}")
    
    # Load the Type B file
    wb_source = load_workbook(filepath, data_only=True)
    ws_source = wb_source['Data']
    
    # Extract data from Type B sheet
    title = ws_source['A1'].value or "Portfolio Report"
    
    # Find month headers (typically row 4)
    month_row = None
    for row in range(1, 10):
        if ws_source[f'B{row}'].value and 'Mar' in str(ws_source[f'B{row}'].value):
            month_row = row
            break
    
    if not month_row:
        print("  [ERROR] Could not find month headers")
        return False
    
    # Extract months and data
    months = []
    for col in range(2, 15):
        cell_val = ws_source.cell(row=month_row, column=col).value
        if cell_val:
            months.append(str(cell_val))
    
    # Extract metrics data
    metrics_data = {}
    for row in range(month_row + 1, ws_source.max_row + 1):
        metric_name = ws_source[f'A{row}'].value
        if metric_name and str(metric_name).strip() and str(metric_name).strip() != '-':
            values = []
            for col in range(2, 2 + len(months)):
                val = ws_source.cell(row=row, column=col).value
                values.append(val)
            metrics_data[str(metric_name).strip()] = values
    
    wb_source.close()
    
    # Calculate KPIs from extracted data
    kpis = calculate_kpis(metrics_data, months, title)
    
    # Create new Type A workbook
    wb_new = Workbook()
    wb_new.remove(wb_new.active)  # Remove default sheet
    
    # Create sheets
    ws_exec = wb_new.create_sheet("Executive Summary")
    ws_monthly = wb_new.create_sheet("Monthly Performance")
    ws_data = wb_new.create_sheet("Data Source")
    
    # Create Executive Summary
    create_executive_summary(ws_exec, title, kpis)
    
    # Create Monthly Performance
    create_monthly_performance(ws_monthly, title, months, metrics_data)
    
    # Copy original data to Data Source
    copy_data_source(ws_data, filepath)
    
    # Apply professional formatting
    format_sheets(wb_new)
    
    # Save the restructured file
    wb_new.save(filepath)
    print(f"  [OK] Restructured to Type A format")
    print(f"  [OK] Created Executive Summary sheet")
    print(f"  [OK] Created Monthly Performance sheet")
    print(f"  [OK] Created Data Source sheet")
    print(f"  [OK] Applied professional formatting")
    print(f"[OK] File saved successfully!\n")
    
    return True


def calculate_kpis(metrics_data, months, title):
    """Calculate Key Performance Indicators from metrics data"""
    
    kpis = {}
    
    # Extract relevant metrics
    portfolio_values = metrics_data.get('Portfolio value', [])
    starting_values = metrics_data.get('At the beginning of the period', portfolio_values[:-1] if portfolio_values else [])
    total_profits = metrics_data.get('Total profit', [])
    profit_percents = metrics_data.get('Total profit, %', [])
    dividends = metrics_data.get('Dividends', [0]*len(months))
    
    if not portfolio_values or len(portfolio_values) < 2:
        return kpis
    
    # Starting and ending values
    kpis['start_value'] = float(starting_values[0]) if starting_values else 0
    kpis['end_value'] = float(portfolio_values[-1]) if portfolio_values else 0
    kpis['growth'] = kpis['end_value'] - kpis['start_value']
    kpis['growth_percent'] = (kpis['growth'] / kpis['start_value'] * 100) if kpis['start_value'] else 0
    
    # Total metrics
    total_profit = sum(float(p) if isinstance(p, (int, float)) else 0 for p in total_profits)
    total_dividends = sum(float(d) if isinstance(d, (int, float)) else 0 for d in dividends)
    kpis['total_profit'] = total_profit
    kpis['total_dividends'] = total_dividends
    kpis['total_gains'] = total_profit - total_dividends
    
    # Monthly statistics
    profit_values = [float(p) if isinstance(p, (int, float)) else 0 for p in total_profits]
    valid_profits = [p for p in profit_values if p is not None and p != 0]
    
    if valid_profits:
        kpis['best_month'] = max(valid_profits)
        kpis['worst_month'] = min(valid_profits)
        kpis['avg_monthly'] = sum(valid_profits) / len(valid_profits)
        kpis['positive_months'] = len([p for p in valid_profits if p > 0])
    else:
        kpis['best_month'] = 0
        kpis['worst_month'] = 0
        kpis['avg_monthly'] = 0
        kpis['positive_months'] = 0
    
    kpis['total_months'] = len(months)
    kpis['months'] = months
    
    return kpis


def create_executive_summary(ws, title, kpis):
    """Create Executive Summary sheet with KPIs and insights"""
    
    # Title
    ws['A1'].value = title
    ws.merge_cells('A1:E1')
    ws.row_dimensions[1].height = 25
    
    # Date
    ws['A2'].value = f"Report Generated: February 7, 2026"
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 8
    
    # KEY PERFORMANCE INDICATORS
    ws['A4'].value = 'KEY PERFORMANCE INDICATORS'
    ws.merge_cells('A4:E4')
    ws.row_dimensions[4].height = 20
    
    # Headers
    ws['A5'].value = 'Metric'
    ws['B5'].value = 'Value'
    ws['C5'].value = 'Details'
    ws.row_dimensions[5].height = 18
    
    # KPI data
    kpi_rows = [
        ('Portfolio Growth', f"${kpis.get('growth', 0):,.2f}", f"{kpis.get('growth_percent', 0):.2f}% increase"),
        ('Starting Value', f"${kpis.get('start_value', 0):,.2f}", f"{kpis.get('months', [])[0] if kpis.get('months') else 'Mar 2025'}"),
        ('Ending Value', f"${kpis.get('end_value', 0):,.2f}", "February 2026"),
        ('Total Profit', f"${kpis.get('total_profit', 0):,.2f}", 'Last 12 months'),
        ('Total Dividends', f"${kpis.get('total_dividends', 0):,.2f}", 'Cumulative'),
        ('Average Monthly Return', f"${kpis.get('avg_monthly', 0):,.2f}", 'Per month average'),
        ('Positive Months', f"{kpis.get('positive_months', 0)} of {kpis.get('total_months', 12)}", f"{(kpis.get('positive_months', 0)/kpis.get('total_months', 12)*100):.0f}% win rate"),
        ('Best Month', f"${kpis.get('best_month', 0):,.2f}", 'Highest profit'),
        ('Worst Month', f"${kpis.get('worst_month', 0):,.2f}", 'Lowest profit'),
    ]
    
    for idx, (metric, value, detail) in enumerate(kpi_rows, start=6):
        ws[f'A{idx}'].value = metric
        ws[f'B{idx}'].value = value
        ws[f'C{idx}'].value = detail
        ws.row_dimensions[idx].height = 16
    
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15


def create_monthly_performance(ws, title, months, metrics_data):
    """Create Monthly Performance sheet with organized monthly data"""
    
    # Title
    ws['A1'].value = 'MONTHLY PERFORMANCE ANALYSIS'
    ws.merge_cells('A1:M1')
    ws.row_dimensions[1].height = 22
    
    ws.row_dimensions[2].height = 8
    
    # Month headers
    ws['A3'].value = 'Period'
    for idx, month in enumerate(months, start=2):
        ws.cell(row=3, column=idx).value = month
    ws.row_dimensions[3].height = 18
    
    # Extract and organize metrics into sections
    row_num = 4
    current_section = None
    
    # Portfolio Values section
    ws[f'A{row_num}'].value = 'Portfolio Value (Start)'
    for col_idx, month in enumerate(months, start=2):
        starting_values = metrics_data.get('At the beginning of the period', [])
        if col_idx - 2 < len(starting_values):
            val = starting_values[col_idx - 2]
            ws.cell(row=row_num, column=col_idx).value = val
    row_num += 1
    
    ws[f'A{row_num}'].value = 'Portfolio Value (End)'
    for col_idx, month in enumerate(months, start=2):
        port_values = metrics_data.get('Portfolio value', [])
        if col_idx - 2 < len(port_values):
            val = port_values[col_idx - 2]
            ws.cell(row=row_num, column=col_idx).value = val
    row_num += 2
    
    # PROFIT METRICS section
    ws[f'A{row_num}'].value = 'PROFIT METRICS'
    ws.merge_cells(f'A{row_num}:M{row_num}')
    ws.row_dimensions[row_num].height = 18
    row_num += 1
    
    profit_metrics = ['Total profit', 'Total profit, %', 'Profit from price change', 
                     'Net profit from sales', 'Dividends']
    
    for metric in profit_metrics:
        if metric in metrics_data:
            ws[f'A{row_num}'].value = metric
            for col_idx, month in enumerate(months, start=2):
                values = metrics_data[metric]
                if col_idx - 2 < len(values):
                    val = values[col_idx - 2]
                    ws.cell(row=row_num, column=col_idx).value = val
            ws.row_dimensions[row_num].height = 16
            row_num += 1
    
    row_num += 1
    
    # TRADING ACTIVITY section
    ws[f'A{row_num}'].value = 'TRADING ACTIVITY'
    ws.merge_cells(f'A{row_num}:M{row_num}')
    ws.row_dimensions[row_num].height = 18
    row_num += 1
    
    trading_metrics = ['Total trades', 'Buy trades', 'Sell trades', 'Total Turnover',
                      'Total purchases', 'Total sales']
    
    for metric in trading_metrics:
        if metric in metrics_data:
            ws[f'A{row_num}'].value = metric
            for col_idx, month in enumerate(months, start=2):
                values = metrics_data[metric]
                if col_idx - 2 < len(values):
                    val = values[col_idx - 2]
                    ws.cell(row=row_num, column=col_idx).value = val
            ws.row_dimensions[row_num].height = 16
            row_num += 1
    
    ws.column_dimensions['A'].width = 25
    for col in range(2, 2 + len(months)):
        ws.column_dimensions[get_column_letter(col)].width = 14


def copy_data_source(ws, filepath):
    """Copy original data to Data Source sheet for reference"""
    
    wb_source = load_workbook(filepath, data_only=True)
    ws_source = wb_source['Data']
    
    ws['A1'].value = 'Original Data Structure'
    ws.row_dimensions[1].height = 20
    
    row_offset = 2
    for row in ws_source.iter_rows(min_row=1, max_row=ws_source.max_row, 
                                     min_col=1, max_col=ws_source.max_column, 
                                     values_only=True):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx).value = value
        row_offset += 1
    
    wb_source.close()


def format_sheets(wb):
    """Apply professional formatting to all sheets"""
    
    # Color scheme
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    subheader_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    metric_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    highlight_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    header_font = Font(bold=True, size=12, color="FFFFFF")
    title_font = Font(bold=True, size=16, color="FFFFFF")
    subheader_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=10)
    
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )
    
    # Format Executive Summary
    ws_exec = wb['Executive Summary']
    ws_exec['A1'].font = title_font
    ws_exec['A1'].fill = header_fill
    ws_exec['A1'].alignment = Alignment(horizontal='left', vertical='center')
    
    ws_exec['A4'].font = subheader_font
    ws_exec['A4'].fill = subheader_fill
    
    for col in ['A', 'B', 'C']:
        ws_exec[col + '5'].font = header_font
        ws_exec[col + '5'].fill = header_fill
        ws_exec[col + '5'].border = thin_border
        ws_exec[col + '5'].alignment = Alignment(horizontal='center', vertical='center')
    
    for row in range(6, 15):
        ws_exec[f'A{row}'].font = bold_font
        ws_exec[f'A{row}'].fill = metric_fill
        ws_exec[f'A{row}'].border = thin_border
        
        ws_exec[f'B{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws_exec[f'B{row}'].border = thin_border
        ws_exec[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws_exec[f'C{row}'].fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        ws_exec[f'C{row}'].border = thin_border
    
    # Format Monthly Performance
    ws_monthly = wb['Monthly Performance']
    ws_monthly['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws_monthly['A1'].fill = header_fill
    
    for col in range(1, 14):
        ws_monthly.cell(row=3, column=col).font = header_font
        ws_monthly.cell(row=3, column=col).fill = header_fill
        ws_monthly.cell(row=3, column=col).border = thin_border
        ws_monthly.cell(row=3, column=col).alignment = Alignment(horizontal='center', vertical='center')
    
    # Format data rows with color coding
    current_section = None
    for row in range(4, ws_monthly.max_row + 1):
        metric_name = ws_monthly[f'A{row}'].value
        if metric_name:
            metric_str = str(metric_name).upper()
            
            if 'PROFIT' in metric_str:
                section_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                current_section = 'profit'
            elif 'TRADING' in metric_str:
                section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                current_section = 'trading'
            elif 'SECTION' not in metric_str:
                if current_section == 'profit':
                    section_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                elif current_section == 'trading':
                    section_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                else:
                    section_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            else:
                section_fill = None
            
            ws_monthly[f'A{row}'].font = bold_font
            ws_monthly[f'A{row}'].fill = metric_fill
            ws_monthly[f'A{row}'].border = thin_border
            
            for col in range(2, 14):
                cell = ws_monthly.cell(row=row, column=col)
                if section_fill:
                    cell.fill = section_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='right', vertical='center')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("\n" + "=" * 70)
        print("TYPE B TO TYPE A RESTRUCTURING TOOL")
        print("Converts single Data sheet to Executive Summary + Monthly Performance")
        print("=" * 70)
        print("\nUsage: python restructure_type_b.py <filename.xlsx> [file2.xlsx ...]")
        print("\nExample:")
        print("  python restructure_type_b.py Portfolio1.xlsx")
        print("  python restructure_type_b.py *.xlsx  (all Type B files)")
        print("=" * 70 + "\n")
        sys.exit(0)
    
    # Get all files to process
    files_to_process = []
    for arg in sys.argv[1:]:
        if '*' in arg:
            import glob
            files_to_process.extend(glob.glob(arg))
        else:
            files_to_process.append(arg)
    
    print("\n" + "=" * 70)
    print(f"Restructuring {len(files_to_process)} file(s) from Type B to Type A...")
    print("=" * 70)
    
    success_count = 0
    error_count = 0
    
    for filename in files_to_process:
        try:
            if restructure_type_b_to_type_a(filename):
                success_count += 1
        except Exception as e:
            print(f"[ERROR] Error processing {filename}: {e}\n")
            error_count += 1
    
    print("=" * 70)
    print(f"COMPLETE: {success_count} file(s) restructured successfully")
    if error_count > 0:
        print(f"ERRORS: {error_count} file(s) failed")
    print("=" * 70 + "\n")
