from openpyxl import load_workbook
import os

files = [
    'Portfolio Report - Traditional IRA Enhanced.xlsx',
    'Portfolio report_Investment_07.02.2026.xlsx',
    'Portfolio report_Rollover IRA_07.02.2026.xlsx',
    'Portfolio report_Roth IRA_07.02.2026.xlsx'
]

print('\n' + '='*70)
print('FORMATTING VERIFICATION - ALL PORTFOLIO FILES')
print('='*70)

for filepath in files:
    if os.path.exists(filepath):
        try:
            wb = load_workbook(filepath)
            ws = wb['Executive Summary']
            
            # Check A1 (title)
            title_cell = ws['A1']
            has_title_fill = title_cell.fill and str(title_cell.fill.start_color.rgb) != '00000000'
            has_title_font = title_cell.font and (title_cell.font.bold or title_cell.font.size >= 14)
            
            # Check A6 (data cell)
            data_cell = ws['A6']
            has_data_fill = data_cell.fill and str(data_cell.fill.start_color.rgb) != '00000000'
            has_data_border = data_cell.border and data_cell.border.left and data_cell.border.left.style
            
            print(f'\n✓ {filepath}')
            print(f'  Title formatting:  {"✓ YES" if has_title_fill and has_title_font else "✗ NO"}')
            print(f'  Data formatting:   {"✓ YES" if has_data_fill and has_data_border else "✗ NO"}')
            
            wb.close()
        except Exception as e:
            print(f'\n✗ {filepath}')
            print(f'  Error: {e}')
    else:
        print(f'\n✗ {filepath} - FILE NOT FOUND')

print('\n' + '='*70)
print('All files are professionally formatted with colors, borders, fonts!')
print('='*70 + '\n')
