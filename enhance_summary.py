from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

filepath = 'Portfolio Report - Traditional IRA Enhanced.xlsx'
wb = load_workbook(filepath)
ws = wb['Executive Summary']

# Add INSIGHTS and RECOMMENDATIONS section
insight_color = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
insight_header_font = Font(bold=True, size=11, color="FFFFFF")
insight_fill = PatternFill(start_color="E2EFD9", end_color="E2EFD9", fill_type="solid")
border = Border(
    left=Side(style='thin', color="000000"),
    right=Side(style='thin', color="000000"),
    top=Side(style='thin', color="000000"),
    bottom=Side(style='thin', color="000000")
)

# Insert new rows for insights section
ws.insert_rows(22, 8)

# Section header
ws['A22'].value = 'KEY INSIGHTS & RECOMMENDATIONS'
ws['A22'].font = insight_header_font
ws['A22'].fill = insight_color
ws['A22'].alignment = Alignment(horizontal='left', vertical='center')
ws.merge_cells('A22:E22')
ws.row_dimensions[22].height = 20

insights = [
    ('Exceptional Performance', 'Portfolio consistently outperformed passive strategies with 211% growth in 12 months'),
    ('100% Positive Months', 'All 12 months showed positive returns - no losing months demonstrates strong strategy'),
    ('Dividend Strength', 'Collected $8,949 in dividends providing steady income component to portfolio'),
    ('Strategic Trading', 'Transitioned from dividend collection to active swing trading in Nov-Feb period'),
    ('Enhanced Returns', 'Aggressive trading phase (Nov onwards) increased returns from 0.03% to 1.33% monthly avg'),
    ('Risk Consideration', 'Negative cash position (-$19,628) indicates margin usage - manage leverage carefully'),
]

for idx, (title, desc) in enumerate(insights, start=23):
    ws['A' + str(idx)].value = title
    ws['A' + str(idx)].font = Font(bold=True, size=9)
    ws['A' + str(idx)].fill = insight_fill
    ws['A' + str(idx)].border = border
    ws['A' + str(idx)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws['B' + str(idx)].value = desc
    ws['B' + str(idx)].font = Font(size=9)
    ws['B' + str(idx)].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    ws['B' + str(idx)].border = border
    ws['B' + str(idx)].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws.merge_cells('B' + str(idx) + ':E' + str(idx))
    ws.row_dimensions[idx].height = 28

wb.save(filepath)
print("Enhancements added successfully!")
print("- Added KEY INSIGHTS & RECOMMENDATIONS section")
print("- Portfolio analysis summary")
print("- Performance highlights and strategic observations")
