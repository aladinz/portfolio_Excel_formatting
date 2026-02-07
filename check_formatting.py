from openpyxl import load_workbook

wb = load_workbook('Portfolio Report - Traditional IRA Enhanced.xlsx')
ws = wb['Executive Summary']

print('Checking formatting in cells A1-A10:')
print()
for row in range(1, 11):
    cell = ws[f'A{row}']
    has_fill = cell.fill and cell.fill.start_color
    has_font = cell.font and (cell.font.bold or cell.font.size != 11)
    has_border = cell.border and cell.border.left and cell.border.left.style
    
    print(f'Row {row}: "{cell.value}"')
    if has_fill:
        print(f'  ✓ Fill color: {cell.fill.start_color.rgb if hasattr(cell.fill.start_color, "rgb") else "indexed"}')
    if has_font:
        print(f'  ✓ Font: size={cell.font.size}, bold={cell.font.bold}')
    if has_border:
        print(f'  ✓ Has border')
    if not (has_fill or has_font or has_border):
        print(f'  ✗ NO FORMATTING')

print()
print('File check complete.')
wb.close()
